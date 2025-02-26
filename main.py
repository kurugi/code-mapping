# 병원코드와 회사코드 연결 (Dictionary 구성)




import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QTableWidget, QTableWidgetItem, QPushButton, QHeaderView
)
from PyQt6.QtGui import QColor

def read_excel_columns_as_records(filename, sheetname=None):
    """
    openpyxl을 사용하여 엑셀 파일의 첫 번째 행(이름)과 두 번째 행(코드)을 읽고,
    각 열을 (이름, 코드) 튜플의 리스트로 반환합니다.
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb[sheetname] if sheetname else wb.active
    names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    codes = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    return list(zip(names, codes))

def merge_records_union(hosp_file, lg_file):
    """
    두 파일의 데이터를 union 방식(모든 이름)으로 병합하여,
    (병원 이름, 병원 코드, LG 이름, LG 코드) 튜플 리스트로 반환합니다.
    매칭되지 않는 항목은 빈 문자열로 채웁니다.
    """
    hosp_records = read_excel_columns_as_records(hosp_file)
    lg_records = read_excel_columns_as_records(lg_file)
    
    filtered_hosp = {}
    for name, code in hosp_records:
        if (name is None or str(name).strip() == "") and (code is None or str(code).strip() == ""):
            continue
        filtered_hosp[str(name).strip()] = code
        
    filtered_lg = {}
    for name, code in lg_records:
        if (name is None or str(name).strip() == "") and (code is None or str(code).strip() == ""):
            continue
        filtered_lg[str(name).strip()] = code
        
    all_names = set(filtered_hosp.keys()).union(set(filtered_lg.keys()))
    merged = []
    for name in sorted(all_names):
        hosp_code = filtered_hosp.get(name, "")
        lg_code = filtered_lg.get(name, "")
        merged.append((name, hosp_code, name, lg_code))
    return merged

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("코드 mapping")
        self.resize(800, 900)

        # 메인 위젯 및 레이아웃 설정
        main_widget = QWidget()
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        # 상단: 병원코드와 LG코드 파일 데이터를 표시하는 TableWidget들을 담는 수평 레이아웃
        table_layout = QHBoxLayout()

        # 병원코드 TableWidget
        self.tableWidgetHospital = QTableWidget()
        self.tableWidgetHospital.setShowGrid(True)
        self.tableWidgetHospital.setAlternatingRowColors(True)
        self.tableWidgetHospital.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tableWidgetHospital.setStyleSheet(
            "QTableWidget { background-color: white; alternate-background-color: lightgray; }"
            "QTableWidget::item:selected { background-color: lightgreen; color: black; }"
        )
        label_hospital = QLabel("병원코드 파일")
        layout_hospital = QVBoxLayout()
        layout_hospital.addWidget(label_hospital)
        layout_hospital.addWidget(self.tableWidgetHospital)
        table_layout.addLayout(layout_hospital)

        # LG코드 TableWidget
        self.tableWidgetLG = QTableWidget()
        self.tableWidgetLG.setShowGrid(True)
        self.tableWidgetLG.setAlternatingRowColors(True)
        self.tableWidgetLG.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tableWidgetLG.setStyleSheet(
            "QTableWidget { background-color: white; alternate-background-color: lightgray; }"
            "QTableWidget::item:selected { background-color: lightgreen; color: black; }"
        )
        label_lg = QLabel("LG코드 파일")
        layout_lg = QVBoxLayout()
        layout_lg.addWidget(label_lg)
        layout_lg.addWidget(self.tableWidgetLG)
        table_layout.addLayout(layout_lg)

        main_layout.addLayout(table_layout)

        # 개별 TableWidget의 헤더 배경색 설정 (예: rgb(200,157,241))
        self.tableWidgetHospital.horizontalHeader().setStyleSheet("background-color: rgb(200,157,241);")
        self.tableWidgetLG.horizontalHeader().setStyleSheet("background-color: rgb(200,157,241);")

        # 중간: '코드연결' 및 '코드해제' 버튼
        button_layout = QHBoxLayout()
        self.connectButton = QPushButton("코드연결")
        self.disconnectButton = QPushButton("코드해제")
        self.connectButton.setFixedHeight(self.connectButton.sizeHint().height() * 2)
        self.disconnectButton.setFixedHeight(self.disconnectButton.sizeHint().height() * 2)
        self.connectButton.setStyleSheet("QPushButton { background-color: lightblue; }")
        self.disconnectButton.setStyleSheet("QPushButton { background-color: lightcoral; }")
        self.connectButton.clicked.connect(self.connect_selected_data)
        self.disconnectButton.clicked.connect(self.disconnect_selected_data)
        button_layout.addWidget(self.connectButton)
        button_layout.addWidget(self.disconnectButton)
        main_layout.addLayout(button_layout)

        # 하단: 연결코드 TableWidget (헤더: 병원 이름, 병원 코드, LG 이름, LG 코드)
        self.connectionTableWidget = QTableWidget()
        self.connectionTableWidget.setShowGrid(True)
        self.connectionTableWidget.setAlternatingRowColors(True)
        self.connectionTableWidget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.connectionTableWidget.setStyleSheet(
            "QTableWidget { background-color: white; alternate-background-color: lightgray; }"
            "QTableWidget::item:selected { background-color: lightgreen; color: black; }"
        )
        self.connectionTableWidget.setColumnCount(4)
        self.connectionTableWidget.setHorizontalHeaderLabels(["병원 이름", "병원 코드", "LG 이름", "LG 코드"])
        self.connectionTableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        # 연결 TableWidget 헤더 배경색은 중간톤 보라색 (#B889EE)
        self.connectionTableWidget.horizontalHeader().setStyleSheet("background-color: #B889EE;")
        main_layout.addWidget(self.connectionTableWidget)

        # 하단: 종료 버튼 (높이 2배, 옅은 주황색)
        self.exitButton = QPushButton("종료")
        self.exitButton.clicked.connect(self.save_and_exit)
        self.exitButton.setFixedHeight(self.exitButton.sizeHint().height() * 2)
        self.exitButton.setStyleSheet("QPushButton { background-color: rgb(214,151,96); }")
        main_layout.addWidget(self.exitButton)

        # 데이터 로드 (병원코드, LG코드 TableWidget 자동 로드)
        self.load_data_to_table('병원코드.xlsx', self.tableWidgetHospital)
        self.load_data_to_table('LG코드.xlsx', self.tableWidgetLG)
        # 프로그램 실행 시 "코드연결.xlsx" 파일이 있으면 연결 TableWidget에 로드
        self.load_connection_from_excel()
        # 하이라이트 갱신: 병원코드와 LG코드 TableWidget 업데이트
        self.update_hosp_table_highlight()
        self.update_lg_table_highlight()

    def connect_selected_data(self):
        """
        병원코드와 LG코드 TableWidget에서 선택한 행의 데이터를 읽어,
        연결 TableWidget에 (병원 이름, 병원 코드, LG 이름, LG 코드) 형태로 추가하고,
        하이라이트를 갱신합니다.
        """
        hosp_items = self.tableWidgetHospital.selectedItems()
        lg_items = self.tableWidgetLG.selectedItems()

        if not hosp_items:
            print("병원코드 TableWidget에서 한 행을 선택하세요.")
            return
        if not lg_items:
            print("LG코드 TableWidget에서 한 행을 선택하세요.")
            return

        hosp_row = hosp_items[0].row()
        lg_row = lg_items[0].row()

        hosp_name = self.tableWidgetHospital.item(hosp_row, 0).text() if self.tableWidgetHospital.item(hosp_row, 0) else ""
        hosp_code = self.tableWidgetHospital.item(hosp_row, 1).text() if self.tableWidgetHospital.item(hosp_row, 1) else ""
        lg_name = self.tableWidgetLG.item(lg_row, 0).text() if self.tableWidgetLG.item(lg_row, 0) else ""
        lg_code = self.tableWidgetLG.item(lg_row, 1).text() if self.tableWidgetLG.item(lg_row, 1) else ""

        row = self.connectionTableWidget.rowCount()
        self.connectionTableWidget.insertRow(row)
        self.connectionTableWidget.setItem(row, 0, QTableWidgetItem(hosp_name))
        self.connectionTableWidget.setItem(row, 1, QTableWidgetItem(hosp_code))
        self.connectionTableWidget.setItem(row, 2, QTableWidgetItem(lg_name))
        self.connectionTableWidget.setItem(row, 3, QTableWidgetItem(lg_code))
        print("선택한 데이터가 연결 TableWidget에 추가되었습니다.")
        self.update_hosp_table_highlight()
        self.update_lg_table_highlight()

    def disconnect_selected_data(self):
        """
        연결 TableWidget에서 선택된 행들을 삭제하고, 하이라이트를 갱신합니다.
        """
        selected_rows = self.connectionTableWidget.selectionModel().selectedRows()
        if not selected_rows:
            print("연결 TableWidget에서 삭제할 행을 선택하세요.")
            return
        rows_to_remove = sorted([index.row() for index in selected_rows], reverse=True)
        for row in rows_to_remove:
            self.connectionTableWidget.removeRow(row)
        print("선택한 행이 연결 TableWidget에서 삭제되었습니다.")
        self.update_hosp_table_highlight()
        self.update_lg_table_highlight()

    def load_data_to_table(self, filename, table_widget):
        records = read_excel_columns_as_records(filename)
        filtered_records = [
            record for record in records
            if not ((record[0] is None or str(record[0]).strip() == "") and 
                    (record[1] is None or str(record[1]).strip() == ""))
        ]
        table_widget.setColumnCount(2)
        table_widget.setHorizontalHeaderLabels(["이름", "코드"])
        table_widget.setRowCount(len(filtered_records))
        for row_idx, (name, code) in enumerate(filtered_records):
            table_widget.setItem(row_idx, 0, QTableWidgetItem(str(name) if name is not None else ""))
            table_widget.setItem(row_idx, 1, QTableWidgetItem(str(code) if code is not None else ""))
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

    def load_connection_from_excel(self):
        filename = "코드연결.xlsx"
        if not os.path.exists(filename):
            print(f"{filename} 파일이 존재하지 않습니다.")
            return
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            print("엑셀에 데이터가 없습니다.")
            return
        data_rows = rows[1:]
        self.connectionTableWidget.clearContents()
        self.connectionTableWidget.setRowCount(len(data_rows))
        for row_index, row in enumerate(data_rows):
            for col_index in range(4):
                value = row[col_index] if col_index < len(row) else ""
                self.connectionTableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value) if value is not None else ""))
        print("코드연결.xlsx 파일에서 데이터를 로드하였습니다.")
        self.update_hosp_table_highlight()
        self.update_lg_table_highlight()

    def save_connection_data_to_excel(self):
        wb = Workbook()
        ws = wb.active
        headers = ["병원 이름", "병원 코드", "LG 이름", "LG 코드"]
        ws.append(headers)
        row_count = self.connectionTableWidget.rowCount()
        col_count = self.connectionTableWidget.columnCount()
        for row in range(row_count):
            row_data = []
            for col in range(col_count):
                item = self.connectionTableWidget.item(row, col)
                row_data.append(item.text() if item is not None else "")
            ws.append(row_data)
        for i in range(1, col_count + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        wb.save("코드연결.xlsx")
        print("코드연결.xlsx 파일이 저장되었습니다.")

    def save_and_exit(self):
        self.save_connection_data_to_excel()
        self.close()

    def update_hosp_table_highlight(self):
        """
        병원코드 TableWidget의 각 행에서, 두 번째 열(병원 코드)이 연결 TableWidget의 두 번째 열에 있으면
        해당 행의 배경색을 노란색으로, 없으면 짝수행은 흰색, 홀수행은 회색으로 설정합니다.
        """
        for row in range(self.tableWidgetHospital.rowCount()):
            code_item = self.tableWidgetHospital.item(row, 1)
            if code_item is None:
                continue
            hosp_code = code_item.text().strip()
            found = False
            for conn_row in range(self.connectionTableWidget.rowCount()):
                conn_item = self.connectionTableWidget.item(conn_row, 1)
                if conn_item and conn_item.text().strip() == hosp_code:
                    found = True
                    break
            for col in range(self.tableWidgetHospital.columnCount()):
                item = self.tableWidgetHospital.item(row, col)
                if item:
                    if found:
                        item.setBackground(QColor("yellow"))
                    else:
                        item.setBackground(QColor("white") if row % 2 == 0 else QColor("lightgray"))

    def update_lg_table_highlight(self):
        """
        LG코드 TableWidget의 각 행에서, 두 번째 열(LG 코드)이 연결 TableWidget의 네 번째 열에 있으면
        해당 행의 배경색을 노란색으로, 없으면 짝수행은 흰색, 홀수행은 회색으로 설정합니다.
        """
        for row in range(self.tableWidgetLG.rowCount()):
            code_item = self.tableWidgetLG.item(row, 1)
            if code_item is None:
                continue
            lg_code = code_item.text().strip()
            found = False
            for conn_row in range(self.connectionTableWidget.rowCount()):
                conn_item = self.connectionTableWidget.item(conn_row, 3)
                if conn_item and conn_item.text().strip() == lg_code:
                    found = True
                    break
            for col in range(self.tableWidgetLG.columnCount()):
                item = self.tableWidgetLG.item(row, col)
                if item:
                    if found:
                        item.setBackground(QColor("yellow"))
                    else:
                        item.setBackground(QColor("white") if row % 2 == 0 else QColor("lightgray"))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
